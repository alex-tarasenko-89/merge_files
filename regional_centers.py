# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
import pandas as pd
import numpy
import openpyxl
import csv


def work_with_file(filename, path, forms_dict):
    if "csv" in filename:
        df = pd.read_csv('{}/{}'.format(path,filename), sep=";")
        df.drop(columns = ["PageNum"], inplace=True)
        apr = filename[:-4]
        if apr in forms_dict:
            forms_dict[apr].append(df, ignore_index=True)
        else:
            forms_dict[apr] = df    
    else:
        xl = openpyxl.load_workbook('{}/{}'.format(path,filename))
        if len(xl.sheetnames) == 1:
            df = pd.DataFrame(xl[0].values)
            h = df.iloc[0,:]
            df.columns = h
            df.drop([0], inplace=True)
            df.drop(columns = ["PageNum"], inplace=True)
            if filename[:-5] in forms_dict:
                forms_dict[filename[:-5]].append(df, ignore_index=True)
            else:
                forms_dict[filename[:-5]] = df
        else:
            for sh in xl.sheetnames:
                df = pd.DataFrame(xl[sh].values)
                h = df.iloc[0,:]
                df.columns = h
                df.drop([0], inplace=True)
                df.drop(columns = ["PageNum"], inplace=True)
                if sh in forms_dict:
                    forms_dict[sh].append(df, ignore_index=True)
                else:
                    forms_dict[sh] = df
        
    return(forms_dict)



def work_with_dir(path):
    forms = {}
    forms_2 = {}
    for filename in os.listdir(path):
        if filename.find('xls') != -1 or filename.find('csv') != -1:
            forms = work_with_file(filename, path, forms)
        else:
            forms_2 = work_with_dir("{}/{}".format(path, filename))
            for key in forms_2.keys():
                if key in forms.keys():
                    forms[key] = forms[key].append(forms_2[key],ignore_index=True)
                else:
                    forms[key] = forms_2[key]
    return(forms)


def check_duplicates(data):
    from collections import Counter
    for key,value in data.items():
        occ = Counter(value['BlankID'])
        n = dict(filter(lambda x: x[1] == 2, occ.items()))
        result = (key,max(occ.keys()),min(occ.keys()), n)
        print(result)
    return(result)

def writing_files(data, path):
    for key, value in data.items():
        value.to_csv("{}/{}.csv".format(path,key), index=False, na_rep="O")
    return("done.")
        
        
path = 'D:/test/bases/test'
f = work_with_dir(path)
writing_files(f, "d:/test")






